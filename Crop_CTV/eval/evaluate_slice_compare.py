"""
Per-slice comparison between baseline and crop_restore predictions.

Only slices with GT foreground are evaluated.
For each (case, z) row, output:
- ID
- current z index
- upper bound z index
- lower bound z index (smaller index)
- normalized z position (smaller means closer to upper bound)
- Dice for baseline and crop_restore
- HD95 for baseline and crop_restore

Output: one Excel file with a summary sheet and one sheet per model.
"""

import math
import os
import re
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import SimpleITK as sitk
from medpy import metric
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


SUMMARY_LABELS = [
    "Upper",
    "0.0-0.1",
    "0.1-0.2",
    "0.2-0.3",
    "0.3-0.4",
    "0.4-0.5",
    "0.5-0.6",
    "0.6-0.7",
    "0.7-0.8",
    "0.8-0.9",
    "0.9-1.0",
    "Lower",
]

NORM_Z_BIN_RANGES = [
    ("0.0-0.1", 0.0, 0.1),
    ("0.1-0.2", 0.1, 0.2),
    ("0.2-0.3", 0.2, 0.3),
    ("0.3-0.4", 0.3, 0.4),
    ("0.4-0.5", 0.4, 0.5),
    ("0.5-0.6", 0.5, 0.6),
    ("0.6-0.7", 0.6, 0.7),
    ("0.7-0.8", 0.7, 0.8),
    ("0.8-0.9", 0.8, 0.9),
    ("0.9-1.0", 0.9, 1.0),
]


def strip_nii_ext(name: str) -> str:
    if name.endswith(".nii.gz"):
        return name[:-7]
    if name.endswith(".nii"):
        return name[:-4]
    return name


def extract_numeric_key(name: str) -> str:
    stem = strip_nii_ext(os.path.basename(name))
    nums = re.findall(r"\d+", stem)
    if not nums:
        raise ValueError(f"No numeric id found in filename: {name}")
    return str(int(nums[-1]))


def list_nii_files(folder: str) -> List[str]:
    return sorted([f for f in os.listdir(folder) if f.endswith(".nii.gz") or f.endswith(".nii")])


def build_numeric_index(folder: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for name in list_nii_files(folder):
        key = extract_numeric_key(name)
        path = os.path.join(folder, name)
        if key in out:
            raise ValueError(
                f"Duplicate numeric id in folder {folder}: id={key}, "
                f"files={os.path.basename(out[key])} and {name}"
            )
        out[key] = path
    return out


def dice_2d(pred_slice: np.ndarray, gt_slice: np.ndarray) -> float:
    pred = pred_slice > 0
    gt = gt_slice > 0
    inter = np.logical_and(pred, gt).sum()
    smooth = 1e-5
    return float((2.0 * inter + smooth) / (pred.sum() + gt.sum() + smooth))


def diag_mm_2d(shape_yx: Tuple[int, int], spacing_xy: Tuple[float, float]) -> float:
    y, x = shape_yx
    sx, sy = spacing_xy
    return float(math.sqrt((y * sy) ** 2 + (x * sx) ** 2))


def hd95_2d_mm(pred_slice: np.ndarray, gt_slice: np.ndarray, spacing_xy: Tuple[float, float]) -> float:
    pred = (pred_slice > 0).astype(np.uint8)
    gt = (gt_slice > 0).astype(np.uint8)

    pred_any = bool(pred.max() > 0)
    gt_any = bool(gt.max() > 0)

    if pred_any and gt_any:
        # medpy expects voxelspacing matching array axis order (y, x)
        return float(metric.binary.hd95(pred, gt, voxelspacing=(spacing_xy[1], spacing_xy[0])))

    # Only GT-positive slices are evaluated; if pred is empty, use 2D diagonal penalty.
    return diag_mm_2d(pred.shape, spacing_xy)


def auto_fit_and_center(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            val_len = len(str(cell.value)) if cell.value is not None else 0
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[letter].width = min(max_len + 2, 45)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


def evaluate_slice_table(
    gt_dir: str,
    baseline_dir: str,
    crop_restore_dir: str,
    out_excel: str,
) -> None:
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    evaluate_slice_sheet(
        wb=wb,
        sheet_name="Slice_Compare",
        gt_dir=gt_dir,
        baseline_dir=baseline_dir,
        crop_restore_dir=crop_restore_dir,
    )

    add_summary_sheet(wb, ["Slice_Compare"])
    wb.save(out_excel)
    print(f"Saved: {out_excel}")


def evaluate_slice_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    gt_dir: str,
    baseline_dir: str,
    crop_restore_dir: str,
) -> int:
    gt_index = build_numeric_index(gt_dir)
    baseline_index = build_numeric_index(baseline_dir)
    crop_index = build_numeric_index(crop_restore_dir)

    columns = [
        "ID",
        "Current Z",
        "Upper Z",
        "Lower Z",
        "Norm Z (0=upper, 1=lower)",
        "Dice baseline",
        "Dice crop_restore",
        "HD95 baseline (mm)",
        "HD95 crop_restore (mm)",
    ]

    rows: List[List[object]] = []

    common_keys = sorted(set(gt_index.keys()) & set(baseline_index.keys()) & set(crop_index.keys()), key=lambda x: int(x))
    missing_baseline = sorted(set(gt_index.keys()) - set(baseline_index.keys()), key=lambda x: int(x))
    missing_crop = sorted(set(gt_index.keys()) - set(crop_index.keys()), key=lambda x: int(x))

    if missing_baseline:
        print(f"[Warn][{sheet_name}] Missing in baseline: {missing_baseline}")
    if missing_crop:
        print(f"[Warn][{sheet_name}] Missing in crop_restore: {missing_crop}")

    for key in common_keys:
        case_id = f"p_{int(key)}"

        gt_img = sitk.ReadImage(gt_index[key])
        baseline_img = sitk.ReadImage(baseline_index[key])
        crop_img = sitk.ReadImage(crop_index[key])

        gt = sitk.GetArrayFromImage(gt_img)
        baseline_pred = sitk.GetArrayFromImage(baseline_img)
        crop_pred = sitk.GetArrayFromImage(crop_img)

        if gt.shape != baseline_pred.shape or gt.shape != crop_pred.shape:
            print(
                f"[Skip][{sheet_name}] Shape mismatch id={key}: "
                f"gt={gt.shape}, baseline={baseline_pred.shape}, crop={crop_pred.shape}"
            )
            continue

        spacing_xyz = gt_img.GetSpacing()  # (sx, sy, sz)
        spacing_xy = (float(spacing_xyz[0]), float(spacing_xyz[1]))

        gt_pos_z = [z for z in range(gt.shape[0]) if np.any(gt[z] > 0)]
        if not gt_pos_z:
            continue

        lower_z = int(min(gt_pos_z))
        upper_z = int(max(gt_pos_z))
        denom = upper_z - lower_z

        # Sort from upper to lower so norm z increases downward.
        for z in sorted(gt_pos_z, reverse=True):
            gt_slice = gt[z]
            baseline_slice = baseline_pred[z]
            crop_slice = crop_pred[z]

            if denom == 0:
                norm_z = 0.0
            else:
                norm_z = float((upper_z - z) / denom)

            dice_baseline = dice_2d(baseline_slice, gt_slice)
            dice_crop = dice_2d(crop_slice, gt_slice)

            hd95_baseline = hd95_2d_mm(baseline_slice, gt_slice, spacing_xy)
            hd95_crop = hd95_2d_mm(crop_slice, gt_slice, spacing_xy)

            rows.append(
                [
                    case_id,
                    int(z),
                    upper_z,
                    lower_z,
                    norm_z,
                    round(dice_baseline, 2),
                    round(dice_crop, 2),
                    round(hd95_baseline, 2),
                    round(hd95_crop, 2),
                ]
            )

    ws = wb.create_sheet(title=sheet_name)

    ws.append(columns)
    for r in rows:
        ws.append(r)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Numeric formatting
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=5).number_format = "0.0000"  # norm z
        for c in range(6, 10):
            ws.cell(row=r, column=c).number_format = "0.00"

    auto_fit_and_center(ws)

    print(f"[{sheet_name}] Rows: {len(rows)}")
    return len(rows)


def mean_skip_nan(values: List[Optional[float]]) -> Optional[float]:
    valid = [v for v in values if v is not None and not math.isnan(v)]
    if not valid:
        return None
    return float(sum(valid) / len(valid))


def get_col_index(header: List[object], name: str) -> int:
    try:
        return header.index(name)
    except ValueError:
        raise ValueError(f"Missing required column: {name}")


def summarize_slice_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> Dict[str, Tuple[Optional[float], Optional[float]]]:
    header = [cell.value for cell in ws[1]]
    id_col = get_col_index(header, "ID")
    norm_col = get_col_index(header, "Norm Z (0=upper, 1=lower)")
    dice_baseline_col = get_col_index(header, "Dice baseline")
    dice_crop_col = get_col_index(header, "Dice crop_restore")

    by_patient: Dict[str, List[Tuple[float, float, float]]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        patient_id = row[id_col]
        norm_z = row[norm_col]
        dice_baseline = row[dice_baseline_col]
        dice_crop = row[dice_crop_col]
        if patient_id is None or norm_z is None:
            continue
        by_patient.setdefault(str(patient_id), []).append(
            (float(norm_z), float(dice_baseline), float(dice_crop))
        )

    per_label_baseline: Dict[str, List[Optional[float]]] = {label: [] for label in SUMMARY_LABELS}
    per_label_crop: Dict[str, List[Optional[float]]] = {label: [] for label in SUMMARY_LABELS}

    for patient_rows in by_patient.values():
        min_norm = min(row[0] for row in patient_rows)
        max_norm = max(row[0] for row in patient_rows)

        upper_rows = [row for row in patient_rows if row[0] == min_norm]
        lower_rows = [row for row in patient_rows if row[0] == max_norm]
        per_label_baseline["Upper"].append(mean_skip_nan([row[1] for row in upper_rows]))
        per_label_crop["Upper"].append(mean_skip_nan([row[2] for row in upper_rows]))
        per_label_baseline["Lower"].append(mean_skip_nan([row[1] for row in lower_rows]))
        per_label_crop["Lower"].append(mean_skip_nan([row[2] for row in lower_rows]))

        for label, lower, upper in NORM_Z_BIN_RANGES:
            if label == "0.9-1.0":
                bin_rows = [row for row in patient_rows if row[0] > lower and row[0] < upper]
            else:
                bin_rows = [row for row in patient_rows if row[0] > lower and row[0] <= upper]

            per_label_baseline[label].append(mean_skip_nan([row[1] for row in bin_rows]))
            per_label_crop[label].append(mean_skip_nan([row[2] for row in bin_rows]))

    summary: Dict[str, Tuple[Optional[float], Optional[float]]] = {}
    for label in SUMMARY_LABELS:
        summary[label] = (
            mean_skip_nan(per_label_baseline[label]),
            mean_skip_nan(per_label_crop[label]),
        )
    return summary


def add_summary_sheet(wb: openpyxl.Workbook, model_sheet_names: List[str]) -> None:
    if "summary" in wb.sheetnames:
        del wb["summary"]

    ws_summary = wb.create_sheet("summary", 0)
    ws_summary.cell(row=1, column=1, value="Label")
    ws_summary.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    summaries = {
        model_name: summarize_slice_sheet(wb[model_name])
        for model_name in model_sheet_names
        if model_name in wb.sheetnames
    }

    col = 2
    for model_name in model_sheet_names:
        if model_name not in summaries:
            continue
        ws_summary.cell(row=1, column=col, value=model_name)
        ws_summary.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws_summary.cell(row=2, column=col, value="Dice_all_mean")
        ws_summary.cell(row=2, column=col + 1, value="Dice_crop_mean")
        col += 2

    for row_idx, label in enumerate(SUMMARY_LABELS, start=3):
        ws_summary.cell(row=row_idx, column=1, value=label)
        col = 2
        for model_name in model_sheet_names:
            if model_name not in summaries:
                continue
            dice_baseline, dice_crop = summaries[model_name][label]
            ws_summary.cell(
                row=row_idx,
                column=col,
                value=None if dice_baseline is None else round(dice_baseline, 2),
            )
            ws_summary.cell(
                row=row_idx,
                column=col + 1,
                value=None if dice_crop is None else round(dice_crop, 2),
            )
            col += 2

    for row in ws_summary.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = Font(bold=True)

    for row in range(3, ws_summary.max_row + 1):
        for col in range(2, ws_summary.max_column + 1):
            ws_summary.cell(row=row, column=col).number_format = "0.00"

    auto_fit_and_center(ws_summary)


def evaluate_slice_workbook(
    gt_dir: str,
    model_dirs: Dict[str, Tuple[str, str]],
    out_excel: str,
) -> None:
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    evaluated = 0
    for sheet_name, (baseline_dir, crop_restore_dir) in model_dirs.items():
        if not os.path.isdir(baseline_dir):
            print(f"[Skip Model] Baseline folder not found: {sheet_name}: {baseline_dir}")
            continue
        if not os.path.isdir(crop_restore_dir):
            print(f"[Skip Model] Crop restore folder not found: {sheet_name}: {crop_restore_dir}")
            continue
        evaluate_slice_sheet(
            wb=wb,
            sheet_name=sheet_name,
            gt_dir=gt_dir,
            baseline_dir=baseline_dir,
            crop_restore_dir=crop_restore_dir,
        )
        evaluated += 1

    if evaluated == 0:
        print("No model evaluated. Excel not generated.")
        return

    add_summary_sheet(wb, list(model_dirs.keys()))
    out_dir = os.path.dirname(out_excel)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(out_excel)
    print(f"Saved: {out_excel}")


if __name__ == "__main__":
    base_dir = r"D:\WUSI\Crop\Eso_83"
    gt_dir = os.path.join(base_dir, "labelsTs")

    model_dirs = {
        "AttentionUNet": (
            os.path.join(base_dir, "AttentionUNet", "AttentionUNet_all_rawpred"),
            os.path.join(base_dir, "AttentionUNet", "AttentionUNet_all_preprocess"),
        ),
        "DDUNet": (
            os.path.join(base_dir, "DDUNet", "DDUNet_all_rawpred"),
            os.path.join(base_dir, "DDUNet", "DDUNet_all_preprocess"),
        ),
        "nnUNet": (
            os.path.join(base_dir, "nnUNet", "nnUNet_all_rawpred"),
            os.path.join(base_dir, "nnUNet", "nnUNet_all_preprocess"),
        ),
    }

    out_excel = r"D:\WUSI\Crop\Analysis\Eval_Results\Eso\Eso_GTslice_Compare.xlsx"

    evaluate_slice_workbook(
        gt_dir=gt_dir,
        model_dirs=model_dirs,
        out_excel=out_excel,
    )
