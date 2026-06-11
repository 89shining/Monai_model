import os
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import openpyxl
import SimpleITK as sitk
from openpyxl.styles import Font

from evaluate_metrics import (
    APL_TOL_MM,
    PM,
    SURFACE_TOL_MM,
    auto_fit_and_center,
    build_numeric_index,
    check_image_consistency,
    compute_2d_metrics_nonempty_slices,
    compute_apl_bidirectional_mm,
    compute_surface_dsc_3d,
    dice_coefficient,
    safe_asd_3d,
    safe_hd95_3d,
    write_model_sheet,
)


@dataclass(frozen=True)
class ModelEvalConfig:
    model_name: str
    pred_root_fullsize: str
    gt_dir: str
    output_excel: str


def default_model_configs() -> List[ModelEvalConfig]:
    return [
        ModelEvalConfig(
            model_name="AttentionUNet",
            pred_root_fullsize="/home/wusi/Project_crop/Data/Eso_83/Networks/AttentionUNet/EsoCTV_Error_FullSize",
            gt_dir="/home/wusi/Project_crop/Data/Eso_83/EsoCTV_All/labelsTs",
            output_excel="/home/wusi/Project_crop/Data/Eso_83/Networks/AttentionUNet/EsoCTV_Error_FullSize/AttentionUNet_Eval_Error_metrics.xlsx",
        ),
        ModelEvalConfig(
            model_name="DDUnet",
            pred_root_fullsize="/home/wusi/Project_crop/Data/Eso_83/Networks/DDUnet/EsoCTV_Error_FullSize",
            gt_dir="/home/wusi/Project_crop/Data/Eso_83/EsoCTV_All/labelsTs",
            output_excel="/home/wusi/Project_crop/Data/Eso_83/Networks/DDUnet/EsoCTV_Error_FullSize/DDUnet_Eval_Error_metrics.xlsx",
        ),
        ModelEvalConfig(
            model_name="Deeplabv3+",
            pred_root_fullsize="/home/wusi/Project_crop/Data/Rectal_146/Networks/Deeplabv3+/RectalCTV_Error_FullSize",
            gt_dir="/home/wusi/Project_crop/Data/Rectal_146/RectalCTV_All/labelsTs",
            output_excel="/home/wusi/Project_crop/Data/Rectal_146/Networks/Deeplabv3+/RectalCTV_Error_FullSize/Deeplabv3+_Eval_Error_metrics.xlsx",
        ),
        ModelEvalConfig(
            model_name="VNet",
            pred_root_fullsize="/home/wusi/Project_crop/Data/Rectal_146/Networks/VNet/RectalCTV_Error_FullSize",
            gt_dir="/home/wusi/Project_crop/Data/Rectal_146/RectalCTV_All/labelsTs",
            output_excel="/home/wusi/Project_crop/Data/Rectal_146/Networks/VNet/RectalCTV_Error_FullSize/VNet_Eval_Error_metrics.xlsx",
        ),
        ModelEvalConfig(
            model_name="nnUNet_Eso",
            pred_root_fullsize="/home/wusi/Project_crop/Data/Eso_83/Networks/nnUNet/EsoCTV_Error_FullSize",
            gt_dir="/home/wusi/Project_crop/Data/Eso_83/EsoCTV_All/labelsTs",
            output_excel="/home/wusi/Project_crop/Data/Eso_83/Networks/nnUNet/EsoCTV_Error_FullSize/nnUNet_Eval_Error_metrics.xlsx",
        ),
        ModelEvalConfig(
            model_name="nnUNet_Rectal",
            pred_root_fullsize="/home/wusi/Project_crop/Data/Rectal_146/Networks/nnUNet/RectalCTV_Error_FullSize",
            gt_dir="/home/wusi/Project_crop/Data/Rectal_146/RectalCTV_All/labelsTs",
            output_excel="/home/wusi/Project_crop/Data/Rectal_146/Networks/nnUNet/RectalCTV_Error_FullSize/nnUNet_Eval_Error_metrics.xlsx",
        ),
    ]


def metric_columns() -> List[str]:
    return [
        "ID",
        "2D Dice",
        "2D HD95 (mm)",
        "3D Dice",
        "3D HD95 (mm)",
        f"Surface DSC ({int(SURFACE_TOL_MM)}mm)",
        "ASD (mm)",
        "APL_add (mm)",
        "APL_del (mm)",
        "APL_total (mm)",
    ]


def list_error_groups(pred_root_fullsize: str) -> List[str]:
    groups = []
    for name in sorted(os.listdir(pred_root_fullsize)):
        group_dir = os.path.join(pred_root_fullsize, name)
        if os.path.isdir(group_dir):
            groups.append(name)
    if not groups:
        raise FileNotFoundError(f"No error-group folders found under {pred_root_fullsize}")
    return groups


def summarize_rows(rows: List[List[object]]) -> Tuple[List[object], List[object]]:
    numeric_matrix = []
    for row in rows:
        values = []
        for value in row[1:]:
            if isinstance(value, (int, float, np.floating)):
                values.append(float(value))
            else:
                values.append(np.nan)
        numeric_matrix.append(values)

    arr = np.asarray(numeric_matrix, dtype=float) if numeric_matrix else np.empty((0, 9))
    mean_row: List[object] = ["Mean"]
    std_row: List[object] = ["STD"]

    if arr.size == 0:
        mean_row.extend([""] * 9)
        std_row.extend([""] * 9)
        return mean_row, std_row

    means = np.nanmean(arr, axis=0)
    stds = np.nanstd(arr, axis=0, ddof=1)
    for idx in range(9):
        mean_row.append(round(float(means[idx]), 2) if not np.isnan(means[idx]) else "")
        std_row.append(round(float(stds[idx]), 2) if not np.isnan(stds[idx]) else "")
    return mean_row, std_row


def evaluate_group(pred_dir: str, gt_index: Dict[str, str]) -> Tuple[List[List[object]], List[object], List[object]]:
    rows: List[List[object]] = []
    pred_index = build_numeric_index(pred_dir)
    keys = sorted(pred_index.keys(), key=lambda value: int(value))

    for key in keys:
        pred_path = pred_index[key]
        case_id = f"p_{int(key)}"
        gt_path = gt_index.get(key, "")

        if not gt_path:
            print(f"[Skip] GT not found for id={key}, pred={os.path.basename(pred_path)}")
            rows.append([case_id, "", "", "", "", "", "", "", "", ""])
            continue

        gt_img = sitk.ReadImage(gt_path)
        pred_img = sitk.ReadImage(pred_path)
        ok, msg = check_image_consistency(gt_img, pred_img)
        if not ok:
            print(f"[Skip] {case_id}: {msg}")
            rows.append([case_id, "", "", "", "", "", "", "", "", ""])
            continue

        gt = sitk.GetArrayFromImage(gt_img)
        pred = sitk.GetArrayFromImage(pred_img)
        spacing_xyz = gt_img.GetSpacing()

        dice2d, hd95_2d = compute_2d_metrics_nonempty_slices(pred, gt, spacing_xyz)
        dice3d = dice_coefficient(pred, gt)
        hd95_3d = safe_hd95_3d(pred, gt, spacing_xyz)
        surface_dsc = compute_surface_dsc_3d(pred, gt, spacing_xyz, tolerance_mm=SURFACE_TOL_MM)
        asd_3d = safe_asd_3d(pred, gt, spacing_xyz)
        apl_add_mm, apl_del_mm, apl_total_mm = compute_apl_bidirectional_mm(
            pred, gt, spacing_xyz, tolerance_mm=APL_TOL_MM
        )

        rows.append(
            [
                case_id,
                round(dice2d, 2) if not np.isnan(dice2d) else "",
                round(hd95_2d, 2) if not np.isnan(hd95_2d) else "",
                round(dice3d, 2) if not np.isnan(dice3d) else "",
                round(hd95_3d, 2) if not np.isnan(hd95_3d) else "",
                round(surface_dsc, 2) if not np.isnan(surface_dsc) else "",
                round(asd_3d, 2) if not np.isnan(asd_3d) else "",
                round(apl_add_mm, 2),
                round(apl_del_mm, 2),
                round(apl_total_mm, 2),
            ]
        )

    mean_row, std_row = summarize_rows(rows)
    return rows, mean_row, std_row


def write_summary_sheet(
    wb: openpyxl.Workbook,
    summary_rows: List[List[object]],
) -> None:
    if "Summary" in wb.sheetnames:
        wb.remove(wb["Summary"])
    ws = wb.create_sheet("Summary", 0)
    columns = ["Error"] + metric_columns()[1:]
    ws.append(columns)
    for row in summary_rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    auto_fit_and_center(ws)


def evaluate_model(config: ModelEvalConfig) -> None:
    print(f"\nEvaluating {config.model_name}")
    os.makedirs(os.path.dirname(config.output_excel), exist_ok=True)
    gt_index = build_numeric_index(config.gt_dir)
    groups = list_error_groups(config.pred_root_fullsize)

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    summary_rows: List[List[object]] = []
    columns = metric_columns()

    for group in groups:
        pred_dir = os.path.join(config.pred_root_fullsize, group)
        rows, mean_row, std_row = evaluate_group(pred_dir, gt_index)
        write_model_sheet(wb, group, columns, rows, mean_row, std_row)

        summary_row = [group]
        for idx in range(1, len(mean_row)):
            mv = mean_row[idx]
            sv = std_row[idx]
            if mv == "" or sv == "":
                summary_row.append("")
            else:
                summary_row.append(f"{float(mv):.2f} {PM} {float(sv):.2f}")
        summary_rows.append(summary_row)

        print(
            f"[{config.model_name}/{group}] "
            f"2D Dice={summary_row[1]} | "
            f"2D HD95={summary_row[2]} | "
            f"3D Dice={summary_row[3]} | "
            f"3D HD95={summary_row[4]} | "
            f"Surface DSC={summary_row[5]} | "
            f"ASD={summary_row[6]} | "
            f"APL_add={summary_row[7]} | "
            f"APL_del={summary_row[8]} | "
            f"APL_total={summary_row[9]}"
        )

    write_summary_sheet(wb, summary_rows)
    wb.save(config.output_excel)
    print(f"Saved: {config.output_excel}")


def main() -> None:
    for config in default_model_configs():
        if not os.path.isdir(config.pred_root_fullsize):
            print(f"[Skip Model] Folder not found: {config.pred_root_fullsize}")
            continue
        if os.path.exists(config.output_excel):
            print(f"[Skip Model] Excel already exists: {config.output_excel}")
            continue
        evaluate_model(config)


if __name__ == "__main__":
    main()
