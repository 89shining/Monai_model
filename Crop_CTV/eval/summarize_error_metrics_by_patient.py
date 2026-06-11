import argparse
import os
import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


KS = ("K1", "K2", "K3")
MODES = ("inward", "outward", "upshift", "downshift")
PM = "\u00B1"


def auto_fit_and_center(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            val_len = len(str(cell.value)) if cell.value is not None else 0
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


@dataclass(frozen=True)
class ModelSummaryConfig:
    model_name: str
    input_excel: str
    output_excel: str


def default_model_configs() -> Dict[str, ModelSummaryConfig]:
    configs = [
        ModelSummaryConfig(
            model_name="AttentionUNet",
            input_excel="/home/wusi/Project_crop/Data/Eso_83/Networks/AttentionUNet/EsoCTV_Error_FullSize/AttentionUNet_Eval_Error_metrics.xlsx",
            output_excel="/home/wusi/Project_crop/Data/Eso_83/Networks/AttentionUNet/EsoCTV_Error_FullSize/AttentionUNet_Eval_Error_patient_summary.xlsx",
        ),
        ModelSummaryConfig(
            model_name="DDUnet",
            input_excel="/home/wusi/Project_crop/Data/Eso_83/Networks/DDUnet/EsoCTV_Error_FullSize/DDUnet_Eval_Error_metrics.xlsx",
            output_excel="/home/wusi/Project_crop/Data/Eso_83/Networks/DDUnet/EsoCTV_Error_FullSize/DDUnet_Eval_Error_patient_summary.xlsx",
        ),
        ModelSummaryConfig(
            model_name="Deeplabv3+",
            input_excel="/home/wusi/Project_crop/Data/Rectal_146/Networks/Deeplabv3+/RectalCTV_Error_FullSize/Deeplabv3+_Eval_Error_metrics.xlsx",
            output_excel="/home/wusi/Project_crop/Data/Rectal_146/Networks/Deeplabv3+/RectalCTV_Error_FullSize/Deeplabv3+_Eval_Error_patient_summary.xlsx",
        ),
        ModelSummaryConfig(
            model_name="VNet",
            input_excel="/home/wusi/Project_crop/Data/Rectal_146/Networks/VNet/RectalCTV_Error_FullSize/VNet_Eval_Error_metrics.xlsx",
            output_excel="/home/wusi/Project_crop/Data/Rectal_146/Networks/VNet/RectalCTV_Error_FullSize/VNet_Eval_Error_patient_summary.xlsx",
        ),
        ModelSummaryConfig(
            model_name="nnUNet_Eso",
            input_excel="/home/wusi/Project_crop/Data/Eso_83/Networks/nnUNet/EsoCTV_Error_FullSize/nnUNet_Eval_Error_metrics.xlsx",
            output_excel="/home/wusi/Project_crop/Data/Eso_83/Networks/nnUNet/EsoCTV_Error_FullSize/nnUNet_Eval_Error_patient_summary.xlsx",
        ),
        ModelSummaryConfig(
            model_name="nnUNet_Rectal",
            input_excel="/home/wusi/Project_crop/Data/Rectal_146/Networks/nnUNet/RectalCTV_Error_FullSize/nnUNet_Eval_Error_metrics.xlsx",
            output_excel="/home/wusi/Project_crop/Data/Rectal_146/Networks/nnUNet/RectalCTV_Error_FullSize/nnUNet_Eval_Error_patient_summary.xlsx",
        ),
    ]
    return {config.model_name: config for config in configs}


def metric_columns() -> List[str]:
    return [
        "2D Dice",
        "2D HD95 (mm)",
        "3D Dice",
        "3D HD95 (mm)",
        "Surface DSC (3mm)",
        "ASD (mm)",
        "APL_add (mm)",
        "APL_del (mm)",
        "APL_total (mm)",
    ]


def parse_error_sheet_name(sheet_name: str) -> Tuple[str, str]:
    match = re.fullmatch(r"(K[123])_(inward|outward|upshift|downshift)", sheet_name)
    if not match:
        raise ValueError(f"Invalid error sheet name: {sheet_name}")
    return match.group(1), match.group(2)


def is_error_sheet(sheet_name: str) -> bool:
    try:
        parse_error_sheet_name(sheet_name)
        return True
    except ValueError:
        return False


def patient_sort_key(patient_id: str) -> Tuple[int, str]:
    nums = re.findall(r"\d+", str(patient_id))
    if nums:
        return int(nums[-1]), str(patient_id)
    return 10**12, str(patient_id)


def as_float(value: object) -> float:
    if isinstance(value, (int, float, np.floating)):
        return float(value)
    if value is None or value == "":
        return float("nan")
    try:
        return float(value)
    except (TypeError, ValueError):
        return float("nan")


def read_patient_metrics(input_excel: str) -> Tuple[List[str], Dict[str, Dict[Tuple[str, str], List[float]]]]:
    wb = openpyxl.load_workbook(input_excel, data_only=True)
    metric_names = metric_columns()
    data: Dict[str, Dict[Tuple[str, str], List[float]]] = {}

    for sheet_name in wb.sheetnames:
        if not is_error_sheet(sheet_name):
            continue

        k_name, mode = parse_error_sheet_name(sheet_name)
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        if len(header) < len(metric_names) + 1:
            raise ValueError(f"Sheet {sheet_name} has too few columns.")

        sheet_metrics = [str(value) for value in header[1 : len(metric_names) + 1]]
        for expected, actual in zip(metric_names, sheet_metrics):
            if expected != actual:
                raise ValueError(
                    f"Unexpected metric column in {sheet_name}: expected {expected!r}, got {actual!r}"
                )

        for row in ws.iter_rows(min_row=2, values_only=True):
            patient_id = row[0]
            if patient_id in (None, "", "Mean", "STD"):
                continue
            values = [as_float(value) for value in row[1 : len(metric_names) + 1]]
            data.setdefault(str(patient_id), {})[(k_name, mode)] = values

    if not data:
        raise ValueError(f"No patient metric rows found in {input_excel}")
    return metric_names, data


def summarize_values(values: List[float]) -> str:
    arr = np.asarray(values, dtype=float)
    arr = arr[~np.isnan(arr)]
    if arr.size < 2:
        return ""
    mean_value = float(np.mean(arr))
    std_value = float(np.std(arr, ddof=1))
    return f"{mean_value:.2f} {PM} {std_value:.2f}"


def mean_for_summary(values: List[float]) -> float:
    arr = np.asarray(values, dtype=float)
    arr = arr[~np.isnan(arr)]
    if arr.size == 0:
        return float("nan")
    return float(np.mean(arr))


def build_by_k_rows(
    metric_names: List[str],
    data: Dict[str, Dict[Tuple[str, str], List[float]]],
) -> Tuple[List[List[object]], Dict[str, List[List[float]]]]:
    rows: List[List[object]] = []
    summary_values = {k_name: [[] for _ in metric_names] for k_name in KS}
    for patient_id in sorted(data.keys(), key=patient_sort_key):
        patient_values = data[patient_id]
        for k_name in KS:
            row: List[object] = [patient_id, k_name]
            for metric_idx in range(len(metric_names)):
                values = [
                    patient_values.get((k_name, mode), [float("nan")] * len(metric_names))[metric_idx]
                    for mode in MODES
                ]
                row.append(summarize_values(values))
                summary_values[k_name][metric_idx].append(mean_for_summary(values))
            rows.append(row)
    return rows, summary_values


def build_by_mode_rows(
    metric_names: List[str],
    data: Dict[str, Dict[Tuple[str, str], List[float]]],
) -> Tuple[List[List[object]], Dict[str, List[List[float]]]]:
    rows: List[List[object]] = []
    summary_values = {mode: [[] for _ in metric_names] for mode in MODES}
    for patient_id in sorted(data.keys(), key=patient_sort_key):
        patient_values = data[patient_id]
        for mode in MODES:
            row = [patient_id, mode]
            for metric_idx in range(len(metric_names)):
                values = [
                    patient_values.get((k_name, mode), [float("nan")] * len(metric_names))[metric_idx]
                    for k_name in KS
                ]
                row.append(summarize_values(values))
                summary_values[mode][metric_idx].append(mean_for_summary(values))
            rows.append(row)
    return rows, summary_values


def safe_sheet_name(name: str, used_names: Optional[set] = None) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name)[:31]
    if used_names is None or cleaned not in used_names:
        return cleaned

    base = cleaned[:28]
    idx = 1
    candidate = f"{base}_{idx}"
    while candidate in used_names:
        idx += 1
        candidate = f"{base}_{idx}"
    return candidate


def write_sheet(wb: openpyxl.Workbook, sheet_name: str, columns: List[str], rows: List[List[object]]) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(columns)
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    auto_fit_and_center(ws)


def build_summary_rows(
    model_values: Dict[str, Dict[str, List[List[float]]]],
    group_order: Tuple[str, ...],
) -> List[List[object]]:
    rows: List[List[object]] = []
    for model_name, group_values in model_values.items():
        for group_name in group_order:
            row: List[object] = [model_name, group_name]
            for values in group_values[group_name]:
                row.append(summarize_values(values))
            rows.append(row)
    return rows


def summarize_model(config: ModelSummaryConfig, overwrite: bool) -> None:
    if not os.path.exists(config.input_excel):
        print(f"[Skip Model] Excel not found: {config.input_excel}")
        return
    if os.path.exists(config.output_excel) and not overwrite:
        print(f"[Skip Model] Output already exists: {config.output_excel}")
        return

    metric_names, data = read_patient_metrics(config.input_excel)
    by_k_rows, _ = build_by_k_rows(metric_names, data)
    by_mode_rows, _ = build_by_mode_rows(metric_names, data)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    write_sheet(
        wb,
        "By_K",
        ["ID", "K"] + metric_names,
        by_k_rows,
    )
    write_sheet(
        wb,
        "By_Type",
        ["ID", "Type"] + metric_names,
        by_mode_rows,
    )

    os.makedirs(os.path.dirname(config.output_excel), exist_ok=True)
    wb.save(config.output_excel)
    print(f"[Saved] {config.model_name}: {config.output_excel}")


def summarize_models_to_one_workbook(
    configs: List[ModelSummaryConfig],
    output_excel: str,
    overwrite: bool,
) -> None:
    existing_configs = [config for config in configs if os.path.exists(config.input_excel)]
    for config in configs:
        if not os.path.exists(config.input_excel):
            print(f"[Skip Model] Excel not found: {config.input_excel}")

    if not existing_configs:
        raise FileNotFoundError("No input Excel files were found.")
    if os.path.exists(output_excel) and not overwrite:
        print(f"[Skip] Output already exists: {output_excel}")
        return

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    detail_sheets: List[Tuple[str, List[str], List[List[object]]]] = []
    summary_by_k: Dict[str, Dict[str, List[List[float]]]] = {}
    summary_by_mode: Dict[str, Dict[str, List[List[float]]]] = {}
    metric_names: List[str] = []
    used_sheet_names = set()

    for config in existing_configs:
        current_metric_names, data = read_patient_metrics(config.input_excel)
        if not metric_names:
            metric_names = current_metric_names
        elif metric_names != current_metric_names:
            raise ValueError(f"Metric columns are inconsistent in {config.input_excel}")

        by_k_rows, by_k_summary_values = build_by_k_rows(metric_names, data)
        by_mode_rows, by_mode_summary_values = build_by_mode_rows(metric_names, data)
        summary_by_k[config.model_name] = by_k_summary_values
        summary_by_mode[config.model_name] = by_mode_summary_values

        by_k_sheet = safe_sheet_name(f"{config.model_name}_By_K", used_sheet_names)
        used_sheet_names.add(by_k_sheet)
        by_type_sheet = safe_sheet_name(f"{config.model_name}_By_Type", used_sheet_names)
        used_sheet_names.add(by_type_sheet)

        detail_sheets.append((by_k_sheet, ["ID", "K"] + metric_names, by_k_rows))
        detail_sheets.append((by_type_sheet, ["ID", "Type"] + metric_names, by_mode_rows))

    write_sheet(
        wb,
        "Summary_By_K",
        ["Model", "K"] + metric_names,
        build_summary_rows(summary_by_k, KS),
    )
    write_sheet(
        wb,
        "Summary_By_Type",
        ["Model", "Type"] + metric_names,
        build_summary_rows(summary_by_mode, MODES),
    )

    for sheet_name, columns, rows in detail_sheets:
        write_sheet(wb, sheet_name, columns, rows)

    output_dir = os.path.dirname(output_excel)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb.save(output_excel)
    print(f"[Saved Combined] {output_excel}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build per-patient K/type summaries from evaluate_error_metrics.py Excel outputs."
    )
    parser.add_argument(
        "--input_excel",
        type=str,
        default="",
        help="Single existing Eval_Error_metrics.xlsx to summarize. If omitted, default model configs are used.",
    )
    parser.add_argument(
        "--output_excel",
        type=str,
        default="",
        help="Output path for --input_excel mode.",
    )
    parser.add_argument(
        "--model_name",
        type=str,
        default="Custom",
        help="Model name used in log output for --input_excel mode.",
    )
    parser.add_argument(
        "--models",
        nargs="+",
        default=[],
        help="Default configured models to include when --input_excel is omitted.",
    )
    parser.add_argument(
        "--separate",
        action="store_true",
        help="Write one output workbook per model instead of one combined workbook.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing output files.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.input_excel:
        if not args.output_excel:
            root, ext = os.path.splitext(args.input_excel)
            args.output_excel = f"{root}_patient_summary{ext}"
        configs = [
            ModelSummaryConfig(
                model_name=args.model_name,
                input_excel=args.input_excel,
                output_excel=args.output_excel,
            )
        ]
        if args.separate:
            for config in configs:
                summarize_model(config, args.overwrite)
        else:
            summarize_models_to_one_workbook(configs, args.output_excel, args.overwrite)
    else:
        default_configs = default_model_configs()
        if args.models:
            unknown = [name for name in args.models if name not in default_configs]
            if unknown:
                raise ValueError(f"Unknown model names: {', '.join(unknown)}")
            configs = [default_configs[name] for name in args.models]
        else:
            configs = list(default_configs.values())

        if args.separate:
            for config in configs:
                summarize_model(config, args.overwrite)
        else:
            if not args.output_excel:
                first_existing = next((config for config in configs if os.path.exists(config.input_excel)), None)
                if first_existing is None:
                    raise FileNotFoundError("No input Excel files were found.")
                args.output_excel = os.path.join(
                    os.path.dirname(first_existing.input_excel),
                    "All_Models_Eval_Error_patient_summary.xlsx",
                )
            summarize_models_to_one_workbook(configs, args.output_excel, args.overwrite)


if __name__ == "__main__":
    main()
